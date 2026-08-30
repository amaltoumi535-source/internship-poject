# app/modules/reporting/exporters/excel_exporter.py
"""Excel format exporter."""

from typing import Any, Dict
from io import BytesIO

from .base import BaseExporter


class ExcelExporter(BaseExporter):
    """Export reports to Excel format."""

    def export(self, data: Dict[str, Any], filename: str) -> bytes:
        """Export data as Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        row = 1
        
        # Title
        if "title" in data:
            title_cell = ws.cell(row=row, column=1)
            title_cell.value = data["title"]
            title_cell.font = Font(bold=True, size=14)
            row += 2
        
        # Summary
        if "summary" in data:
            ws.cell(row=row, column=1, value="Summary:")
            row += 1
            ws.cell(row=row, column=1, value=data["summary"])
            ws.merge_cells(f"A{row}:D{row}")
            row += 2
        
        # Metadata
        if "metadata" in data:
            ws.cell(row=row, column=1, value="Metadata").font = Font(bold=True)
            row += 1
            for key, value in data["metadata"].items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            row += 1
        
        # Results
        if "results" in data:
            ws.cell(row=row, column=1, value="Results").font = Font(bold=True)
            row += 1
            results = data["results"]
            if isinstance(results, dict):
                for key, value in results.items():
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
            else:
                ws.cell(row=row, column=1, value=str(results))
            row += 2
        
        # Chunks
        if "chunks" in data and data["chunks"]:
            ws.cell(row=row, column=1, value="Chunks").font = Font(bold=True)
            row += 1
            for i, chunk in enumerate(data["chunks"], 1):
                chunk_text = chunk.get("text", str(chunk)) if isinstance(chunk, dict) else str(chunk)
                ws.cell(row=row, column=1, value=f"Chunk {i}").font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value=chunk_text)
                ws.merge_cells(f"A{row}:D{row}")
                row += 1
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def get_extension(self) -> str:
        return "xlsx"

    def get_mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
