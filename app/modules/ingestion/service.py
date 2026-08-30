from pathlib import Path
import shutil
import tempfile
import mimetypes
from typing import Optional, Dict, Any, List

from .parsers.pdf import PDFParser
from .parsers.ocr import OCRParser
from .parsers.docx import DOCXParser
from .parsers.chunker import SemanticChunker

# Optional import for PDF->image conversion for scanned PDFs
try:
    from pdf2image import convert_from_path
    _HAS_PDF2IMAGE = True
except Exception:
    _HAS_PDF2IMAGE = False


class IngestionService:
    def __init__(self, chunk_size: int = 500, chunk_overlap: int = 50):
        self.chunker = SemanticChunker(chunk_size=chunk_size, chunk_overlap=chunk_overlap)

    def _save_upload(self, upload_file: Any) -> Path:
        """Save Starlette UploadFile to a temp Path and return it."""
        suffix = Path(getattr(upload_file, "filename", "")).suffix or ""
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tf:
            upload_path = Path(tf.name)
            # upload_file.file is a SpooledTemporaryFile / file-like
            upload_file.file.seek(0)
            shutil.copyfileobj(upload_file.file, tf)
        return upload_path

    def ingest(self, upload_file: Any, lang: str = "multi") -> Dict[str, Any]:
        """
        Ingest an uploaded file-like object (FastAPI UploadFile).
        Returns dict with metadata, extracted text/pages, and chunks.
        """
        tmp_path = self._save_upload(upload_file)
        try:
            ext = tmp_path.suffix.lower()
            result: Dict[str, Any] = {
                "filename": getattr(upload_file, "filename", str(tmp_path.name)),
                "parser": None,
                "extraction": None,
                "chunks": []
            }

            # PDFs
            if ext in [".pdf"]:
                # Decide whether native text PDF or scanned
                is_text = PDFParser.is_text_pdf(tmp_path)
                if is_text:
                    extraction = PDFParser.extract(tmp_path)
                    result["parser"] = "pdfplumber"
                else:
                    # scanned PDF -> OCR on pages
                    if _HAS_PDF2IMAGE:
                        images = convert_from_path(str(tmp_path))
                        pages: List[Dict[str, Any]] = []
                        full_texts: List[str] = []
                        for i, img in enumerate(images, start=1):
                            # save PIL image to temp file for OCRParser preprocessing
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as img_tf:
                                img.save(img_tf, format="PNG")
                                img_path = Path(img_tf.name)
                            try:
                                ocr_out = OCRParser.extract(img_path, lang=lang)
                                pages.append({
                                    "page_num": i,
                                    "text": ocr_out.get("text", ""),
                                    "confidence": ocr_out.get("confidence", 0.0)
                                })
                                if ocr_out.get("text"):
                                    full_texts.append(ocr_out.get("text"))
                            finally:
                                try:
                                    img_path.unlink()
                                except Exception:
                                    pass

                        extraction = {
                            "text": "\n\n".join(full_texts),
                            "metadata": {"total_pages": len(pages), "parser": "tesseract_pdf_images"},
                            "pages": pages
                        }
                        result["parser"] = "tesseract_pdf_images"
                    else:
                        raise RuntimeError(
                            "PDF looks scanned but pdf2image/poppler not available to convert pages to images."
                        )

            # DOCX
            elif ext in [".docx"]:
                extraction = DOCXParser.extract(tmp_path)
                result["parser"] = "python-docx"

            # Excel
            elif ext in [".xlsx", ".xls"]:
                try:
                    import openpyxl
                except ImportError:
                    raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")

                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                text_parts = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text_parts.append(f"--- Sheet: {sheet_name} ---")
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " | ".join(str(cell) for cell in row if cell is not None)
                        if row_text.strip():
                            text_parts.append(row_text)

                full_text = "\n".join(text_parts)
                extraction = {
                    "text": full_text,
                    "metadata": {"parser": "openpyxl", "sheet_count": len(wb.sheetnames)},
                    "pages": [{"page_num": 1, "text": full_text, "confidence": 1.0}]
                }
                result["parser"] = "openpyxl"

            # Images (common image types)
            elif ext in [".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".gif", ".webp"]:
                extraction = OCRParser.extract(tmp_path, lang=lang)
                # normalize pages format
                extraction = {
                    "text": extraction.get("text", ""),
                    "metadata": extraction.get("metadata", {}),
                    "pages": [{"page_num": 1, "text": extraction.get("text", ""), "confidence": extraction.get("confidence", 0.0)}]
                }
                result["parser"] = "tesseract_image"

            else:
                # Fallback: try PDFParser, else error
                # Try to guess mime and attempt text extraction
                mime, _ = mimetypes.guess_type(tmp_path.name)
                if mime and mime.startswith("text"):
                    text = tmp_path.read_text(encoding="utf-8", errors="ignore")
                    extraction = {"text": text, "metadata": {"parser": "plain_text"}, "pages": [{"page_num": 1, "text": text, "confidence": 1.0}]}
                    result["parser"] = "plain_text"
                else:
                    raise ValueError(f"Unsupported file type: {ext}")

            result["extraction"] = extraction

            # Capture size / mime / page count into extraction metadata
            file_size = tmp_path.stat().st_size
            mime_type, _ = mimetypes.guess_type(result["filename"])
            page_count = len(extraction.get("pages", [])) or extraction.get("metadata", {}).get("total_pages")

            extraction.setdefault("metadata", {})
            extraction["metadata"]["size_bytes"] = file_size
            extraction["metadata"]["mime"] = mime_type
            extraction["metadata"]["page_count"] = page_count

            # Chunking
            text_to_chunk = extraction.get("text", "")
            chunks: List[Dict[str, Any]] = []
            if text_to_chunk and text_to_chunk.strip():
                chunks = self.chunker.chunk(text_to_chunk, metadata={
                    "source_filename": getattr(upload_file, "filename", str(tmp_path.name)),
                    **extraction.get("metadata", {})
                })
            result["chunks"] = chunks
            result["chunks_count"] = len(chunks)
            return result

        finally:
            try:
                tmp_path.unlink()
            except Exception:
                pass